"""
Compare Release 1.23.0 C1 vs C2 Excel files.
Find tasks in C2 that are NEW (not in C1) — i.e., deduplicate.
Matches by ClickUp URL task ID.
"""
import openpyxl

C1_PATH = r"C:\Users\ReemaSingh\Downloads\Release 1.23.0 and C1.xlsx"
C2_PATH = r"C:\Users\ReemaSingh\Downloads\Release 1.23.0 -C2.xlsx"


def read_all_urls(path):
    """Read ALL URLs from ALL sheets in the workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    urls = set()
    tasks_by_url = {}
    
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            # Find the URL column (look for clickup.com)
            url = ""
            task_name = ""
            for i, cell in enumerate(row):
                val = str(cell or "").strip()
                if "clickup.com" in val:
                    url = val
                elif i <= 1 and val and not val.replace(".", "").isdigit():
                    task_name = val
            
            if url:
                tid = url.rstrip("/").split("/")[-1]
                urls.add(tid)
                tasks_by_url[tid] = {"name": task_name, "url": url, "sheet": name}
    
    return urls, tasks_by_url


def read_c2_tasks(path):
    """Read tasks from C2 Tasks sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Tasks"]
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            url = str(row[2] or "").strip()
            tid = url.rstrip("/").split("/")[-1] if url else ""
            tasks.append({
                "num": str(row[0]).strip(),
                "name": str(row[1]).strip(),
                "url": url,
                "assignee": str(row[3] or "").strip(),
                "tid": tid,
            })
    return tasks


# Read C1 — all sheets (FeatureList, Closed Bugs, QASpillover, Openbug)
print(f"Reading C1: {C1_PATH}")
c1_urls, c1_tasks_by_url = read_all_urls(C1_PATH)
print(f"  Total unique task IDs across all C1 sheets: {len(c1_urls)}")

# Read C2 Tasks
print(f"\nReading C2: {C2_PATH}")
c2_tasks = read_c2_tasks(C2_PATH)
print(f"  C2 Tasks: {len(c2_tasks)}")

# Deduplicate
new_in_c2 = []
common = []

for t in c2_tasks:
    if t["tid"] and t["tid"] in c1_urls:
        common.append(t)
    else:
        new_in_c2.append(t)

print(f"\n{'='*70}")
print(f"DEDUPLICATION RESULTS")
print(f"{'='*70}")
print(f"  C1 unique tasks (all sheets): {len(c1_urls)}")
print(f"  C2 Tasks sheet:               {len(c2_tasks)}")
print(f"  Common (already in C1):        {len(common)}")
print(f"  NEW in C2 only:                {len(new_in_c2)}")

print(f"\n{'='*70}")
print(f"TASKS ALREADY IN C1 (carried over, {len(common)}):")
print(f"{'='*70}")
for i, t in enumerate(common, 1):
    c1_info = c1_tasks_by_url.get(t["tid"], {})
    print(f"  {i:3}. [{t['num']}] {t['name']}")
    print(f"       C1 sheet: {c1_info.get('sheet', '?')}")

print(f"\n{'='*70}")
print(f"NEW TASKS IN C2 ONLY ({len(new_in_c2)}):")
print(f"{'='*70}")
for i, t in enumerate(new_in_c2, 1):
    print(f"  {i:3}. [{t['num']}] {t['name']}")
    if t['assignee']:
        print(f"       Assignee: {t['assignee']}")
