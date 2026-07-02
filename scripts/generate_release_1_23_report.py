"""
Generate Release Report for Release 1.23.0-C2 from the ClickUp task Excel.
Reads the actual Excel structure:
  - Tasks sheet: delivered features/fixes
  - QA Spillover: items carried to next cycle
  - Open Bugs for C3: known open issues
  - Closed Bugs: bugs fixed in this release
"""

import openpyxl
from jinja2 import Template
from datetime import datetime
import os

EXCEL_PATH = r"C:\Users\ReemaSingh\Downloads\Release 1.23.0 -C2.xlsx"
OUTPUT_DIR = r"C:\Users\ReemaSingh\Downloads"


def read_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # Tasks
    tasks = []
    ws = wb["Tasks"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            tasks.append({
                "num": str(row[0]).strip(),
                "task": str(row[1]).strip(),
                "url": str(row[2] or "").strip(),
                "assignee": str(row[3] or "").strip(),
            })

    # QA Spillover
    spillover = []
    ws = wb["QA Spillover"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            spillover.append({
                "num": str(row[0]).strip(),
                "task": str(row[1]).strip(),
                "url": str(row[2] or "").strip(),
                "comment": str(row[3] or "").strip(),
            })

    # Open Bugs for C3
    open_bugs = []
    ws = wb["Open Bugs for C3"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            open_bugs.append({
                "num": str(row[0]).strip(),
                "description": str(row[1]).strip(),
                "url": str(row[2] or "").strip(),
                "assignee": str(row[3] or "").strip(),
            })

    # Closed Bugs
    closed_bugs = []
    ws = wb["Closed Bugs"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            closed_bugs.append({
                "num": str(row[0]).strip(),
                "description": str(row[1]).strip(),
                "url": str(row[2] or "").strip(),
                "assignee": str(row[3] or "").strip(),
            })

    return tasks, spillover, open_bugs, closed_bugs


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Release Report — FacctList v1.23.0-C2</title>
<style>
:root { --primary: #1a3a5c; --success: #2e7d32; --danger: #c62828; --warning: #e65100; --bg: #f9fafb; }
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: 'Segoe UI', -apple-system, sans-serif; color: #333; background: var(--bg); line-height: 1.6; }
.container { max-width: 1200px; margin: 30px auto; background: white; padding: 50px 60px; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }
.header { border-bottom: 3px solid var(--primary); padding-bottom: 20px; margin-bottom: 30px; }
.header h1 { color: var(--primary); font-size: 26px; }
.header .subtitle { color: #666; font-size: 14px; margin-top: 5px; }
h2 { color: var(--primary); font-size: 18px; margin: 30px 0 15px 0; padding-bottom: 8px; border-bottom: 2px solid #eef2f7; }
h3 { color: #444; font-size: 15px; margin: 15px 0 10px 0; }

.cards { display: flex; gap: 16px; margin: 20px 0 30px 0; flex-wrap: wrap; }
.card { flex: 1; min-width: 140px; padding: 20px 16px; border-radius: 8px; text-align: center; border: 1px solid #e8e8e8; }
.card .num { font-size: 32px; font-weight: 700; }
.card .label { font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: 0.8px; margin-top: 4px; }
.card-delivered { background: #e8f5e9; border-color: #a5d6a7; }
.card-delivered .num { color: var(--success); }
.card-bugs-fixed { background: #e3f2fd; border-color: #90caf9; }
.card-bugs-fixed .num { color: #1565c0; }
.card-open { background: #fff3e0; border-color: #ffcc80; }
.card-open .num { color: var(--warning); }
.card-spillover { background: #f3e5f5; border-color: #ce93d8; }
.card-spillover .num { color: #7b1fa2; }

table { width: 100%; border-collapse: collapse; margin: 12px 0 20px 0; font-size: 12.5px; }
th { background: var(--primary); color: white; padding: 10px 10px; text-align: left; font-weight: 600; font-size: 11.5px; }
td { padding: 8px 10px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }
tr:hover { background: #f8fafc; }
a { color: #1565c0; text-decoration: none; }
a:hover { text-decoration: underline; }

.tag { display: inline-block; padding: 2px 7px; border-radius: 3px; font-size: 10px; font-weight: 600; }
.tag-feature { background: #e3f2fd; color: #1565c0; }
.tag-bug { background: #fce4ec; color: #c62828; }
.tag-infra { background: #eceff1; color: #455a64; }
.tag-recon { background: #fff8e1; color: #f57f17; }
.tag-ui { background: #f3e5f5; color: #6a1b9a; }

.spillover-comment { font-size: 11px; color: #666; font-style: italic; }
.footer { margin-top: 40px; padding-top: 15px; border-top: 1px solid #e8e8e8; font-size: 11px; color: #aaa; text-align: center; }

.sign-off { margin-top: 25px; }
.sign-off td { padding: 14px 12px; min-width: 100px; border: 1px solid #e8e8e8; }

@media print {
    body { background: white; }
    .container { box-shadow: none; margin: 0; padding: 20px; }
    table { page-break-inside: avoid; }
}
</style>
</head>
<body>
<div class="container">

<div class="header">
    <h1>FacctList — Release Report v1.23.0-C2</h1>
    <p class="subtitle">Release Cycle 2 | June 2026 | QA Team</p>
</div>

<!-- SUMMARY CARDS -->
<h2>1. Executive Summary</h2>
<div class="cards">
    <div class="card card-delivered"><div class="num">{{ tasks|length }}</div><div class="label">Tasks Delivered</div></div>
    <div class="card card-bugs-fixed"><div class="num">{{ closed_bugs|length }}</div><div class="label">Bugs Fixed</div></div>
    <div class="card card-open"><div class="num">{{ open_bugs|length }}</div><div class="label">Open Bugs (C3)</div></div>
    <div class="card card-spillover"><div class="num">{{ spillover|length }}</div><div class="label">QA Spillover</div></div>
</div>

<p>Release <strong>v1.23.0-C2</strong> includes {{ tasks|length }} delivered tasks covering new features,
bug fixes, infrastructure changes, and UI enhancements. {{ closed_bugs|length }} bugs were closed during this cycle.
{{ open_bugs|length }} bugs remain open and are targeted for C3. {{ spillover|length }} items have spilled over to the next QA cycle.</p>

<!-- NEW FEATURES -->
<h2>2. Tasks Delivered</h2>
<table>
<thead><tr><th>#</th><th>Task</th><th>Assignee</th><th>Link</th></tr></thead>
<tbody>
{% for t in tasks %}
<tr>
    <td>{{ t.num }}</td>
    <td>{{ t.task }}</td>
    <td>{{ t.assignee }}</td>
    <td>{% if t.url %}<a href="{{ t.url }}" target="_blank">ClickUp</a>{% endif %}</td>
</tr>
{% endfor %}
</tbody>
</table>

<!-- CLOSED BUGS -->
<h2>3. Bugs Fixed ({{ closed_bugs|length }})</h2>
<table>
<thead><tr><th>#</th><th>Description</th><th>Assignee</th><th>Link</th></tr></thead>
<tbody>
{% for b in closed_bugs %}
<tr>
    <td>{{ b.num }}</td>
    <td>{{ b.description }}</td>
    <td>{{ b.assignee }}</td>
    <td>{% if b.url %}<a href="{{ b.url }}" target="_blank">ClickUp</a>{% endif %}</td>
</tr>
{% endfor %}
</tbody>
</table>

<!-- OPEN BUGS -->
<h2>4. Open Bugs for C3 ({{ open_bugs|length }})</h2>
<table>
<thead><tr><th>#</th><th>Description</th><th>Assignee</th><th>Link</th></tr></thead>
<tbody>
{% for b in open_bugs %}
<tr>
    <td>{{ b.num }}</td>
    <td>{{ b.description }}</td>
    <td>{{ b.assignee }}</td>
    <td>{% if b.url %}<a href="{{ b.url }}" target="_blank">ClickUp</a>{% endif %}</td>
</tr>
{% endfor %}
</tbody>
</table>

<!-- QA SPILLOVER -->
<h2>5. QA Spillover ({{ spillover|length }})</h2>
<table>
<thead><tr><th>#</th><th>Task</th><th>Comment</th><th>Link</th></tr></thead>
<tbody>
{% for s in spillover %}
<tr>
    <td>{{ s.num }}</td>
    <td>{{ s.task }}</td>
    <td><span class="spillover-comment">{{ s.comment }}</span></td>
    <td>{% if s.url %}<a href="{{ s.url }}" target="_blank">ClickUp</a>{% endif %}</td>
</tr>
{% endfor %}
</tbody>
</table>

<!-- SIGN-OFF -->
<h2>6. Sign-Off</h2>
<div class="sign-off">
<table>
<thead><tr><th>Role</th><th>Name</th><th>Date</th><th>Signature</th></tr></thead>
<tbody>
<tr><td>QA Lead</td><td></td><td></td><td></td></tr>
<tr><td>Dev Lead</td><td></td><td></td><td></td></tr>
<tr><td>Product Owner</td><td></td><td></td><td></td></tr>
<tr><td>Release Manager</td><td></td><td></td><td></td></tr>
</tbody>
</table>
</div>

<div class="footer">
    Generated on {{ generated_at }} | FacctList Release v1.23.0-C2 | Confidential — Internal Use
</div>

</div>
</body>
</html>"""


def main():
    print(f"📄 Reading: {EXCEL_PATH}")
    tasks, spillover, open_bugs, closed_bugs = read_excel(EXCEL_PATH)
    print(f"   Tasks: {len(tasks)} | Closed Bugs: {len(closed_bugs)} | Open Bugs: {len(open_bugs)} | Spillover: {len(spillover)}")

    template = Template(HTML_TEMPLATE)
    html = template.render(
        tasks=tasks,
        spillover=spillover,
        open_bugs=open_bugs,
        closed_bugs=closed_bugs,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )

    html_path = os.path.join(OUTPUT_DIR, "Release_1.23.0-C2_QA_Report.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"\n📊 HTML Report: {html_path}")

    # Try PDF
    try:
        from weasyprint import HTML as WPHTML
        pdf_path = os.path.join(OUTPUT_DIR, "Release_1.23.0-C2_QA_Report.pdf")
        WPHTML(string=html).write_pdf(pdf_path)
        print(f"📄 PDF Report:  {pdf_path}")
    except ImportError:
        print("   PDF: Skipped (pip install weasyprint for PDF)")

    print("\n✅ Done!")


if __name__ == "__main__":
    main()
