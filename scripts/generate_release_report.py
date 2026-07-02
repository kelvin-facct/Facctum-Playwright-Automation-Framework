"""
=============================================================================
 FacctList Release Report Generator
=============================================================================
 Reusable script to generate professional HTML + PDF release reports
 from an Excel workbook containing tasks/test data for any release.

 EXCEL FORMAT:
 =============
 The Excel workbook should have these sheets:

 Sheet: "Release Info" (required)
   Key-value pairs in columns A & B:
   | Product         | FacctList                           |
   | Release         | v1.3.0                              |
   | Date            | June 2026                           |
   | Sprint          | Sprint 24                           |
   | Prepared By     | QA Team                             |
   | Environment     | QA - https://qa-saas.facctum.com    |
   | Browser         | Chromium (latest)                   |
   | Framework       | Playwright + Cucumber.js            |

 Sheet: "Tasks" (required)
   Columns: Task ID | Module | Feature | Description | Type | Priority | Status | Comments
   - Type: "New Feature", "Bug Fix", "Enhancement", "Regression", "Technical Debt"
   - Priority: P1, P2, P3
   - Status: Pass, Fail, Blocked, Skipped

 Sheet: "Defects" (optional)
   Columns: Defect ID | Severity | Module | Description | Status | JIRA Link

 Sheet: "Environment" (optional)
   Columns: Parameter | Value
   (e.g., URL, DB, Users, Browser config)

 Sheet: "Risks" (optional)
   Columns: Type | Description | Mitigation

 USAGE:
 ======
   python generate_release_report.py <path_to_excel.xlsx>
   python generate_release_report.py <path_to_excel.xlsx> --output C:\\Reports
   python generate_release_report.py --sample   (creates a sample Excel template)

 OUTPUT:
   - <release>_release_report.html   (styled report)
   - <release>_release_report.pdf    (if weasyprint/pdfkit installed)

 REQUIREMENTS:
   pip install openpyxl jinja2
   pip install weasyprint   (optional, for PDF)
=============================================================================
"""

import sys
import os
import argparse
from datetime import datetime
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("ERROR: Install openpyxl → pip install openpyxl")

try:
    from jinja2 import Template
except ImportError:
    sys.exit("ERROR: Install jinja2 → pip install jinja2")


# =============================================================================
# HTML TEMPLATE
# =============================================================================
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Release Report - {{ info.product }} {{ info.release }}</title>
<style>
:root {
    --primary: #1a3a5c;
    --primary-light: #2c5f8a;
    --success: #2e7d32;
    --danger: #c62828;
    --warning: #f57c00;
    --bg: #f9fafb;
    --card-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: 'Segoe UI', -apple-system, sans-serif; color: #333; background: var(--bg); line-height: 1.6; }
.container { max-width: 1100px; margin: 30px auto; background: white; padding: 50px 60px; border-radius: 10px; box-shadow: var(--card-shadow); }
.header { border-bottom: 3px solid var(--primary); padding-bottom: 20px; margin-bottom: 35px; }
.header h1 { color: var(--primary); font-size: 26px; margin-bottom: 4px; }
.header .subtitle { color: #666; font-size: 14px; }
.header .badge-release { display: inline-block; background: var(--primary); color: white; padding: 3px 12px; border-radius: 4px; font-size: 12px; font-weight: 600; margin-left: 10px; }
h2 { color: var(--primary); font-size: 18px; margin: 35px 0 15px 0; padding-bottom: 8px; border-bottom: 2px solid #eef2f7; }
h3 { color: #444; font-size: 15px; margin: 18px 0 10px 0; }

/* Summary Cards */
.cards { display: flex; gap: 16px; margin: 20px 0 30px 0; flex-wrap: wrap; }
.card { flex: 1; min-width: 130px; padding: 18px 14px; border-radius: 8px; text-align: center; border: 1px solid #e8e8e8; }
.card .num { font-size: 30px; font-weight: 700; color: var(--primary); }
.card .label { font-size: 11px; color: #888; text-transform: uppercase; letter-spacing: 0.8px; margin-top: 4px; }
.card-pass { background: #f0faf0; border-color: #c8e6c9; }
.card-pass .num { color: var(--success); }
.card-fail { background: #fff5f5; border-color: #ffcdd2; }
.card-fail .num { color: var(--danger); }
.card-rate { background: #f5f0ff; border-color: #d1c4e9; }
.card-rate .num { color: #5e35b1; }
.card-new { background: #e8f4fd; border-color: #b3d9f2; }
.card-new .num { color: #0277bd; }

/* Tables */
table { width: 100%; border-collapse: collapse; margin: 12px 0 20px 0; font-size: 13px; }
th { background: var(--primary); color: white; padding: 10px 12px; text-align: left; font-weight: 600; font-size: 12px; }
td { padding: 9px 12px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }
tr:hover { background: #f8fafc; }
.meta-table td { border: 1px solid #e8e8e8; }
.meta-table td:first-child { font-weight: 600; background: #f5f8fa; width: 180px; }

/* Badges */
.badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }
.badge-p1 { background: #ffcdd2; color: #b71c1c; }
.badge-p2 { background: #fff3e0; color: #e65100; }
.badge-p3 { background: #e8f5e9; color: #2e7d32; }
.badge-new { background: #e3f2fd; color: #1565c0; }
.badge-bug { background: #fce4ec; color: #c62828; }
.badge-enh { background: #f3e5f5; color: #6a1b9a; }
.badge-reg { background: #fff8e1; color: #f57f17; }
.badge-tech { background: #eceff1; color: #455a64; }
.status-pass { color: var(--success); font-weight: 600; }
.status-fail { color: var(--danger); font-weight: 600; }
.status-blocked { color: var(--warning); font-weight: 600; }

/* Features List */
.feature-list { margin: 10px 0; padding-left: 0; }
.feature-list li { list-style: none; padding: 8px 12px; margin: 4px 0; background: #f8fafc; border-left: 3px solid var(--primary-light); border-radius: 4px; }
.feature-list li .feat-module { font-weight: 600; color: var(--primary); }
.feature-list li .feat-desc { color: #555; }

/* Footer */
.footer { margin-top: 40px; padding-top: 15px; border-top: 1px solid #e8e8e8; font-size: 11px; color: #aaa; text-align: center; }
.sign-off { margin-top: 25px; }
.sign-off td { padding: 14px 12px; min-width: 100px; border: 1px solid #e8e8e8; }

/* No-defects */
.no-defects { color: var(--success); font-weight: 600; padding: 10px; background: #f0faf0; border-radius: 6px; border: 1px solid #c8e6c9; }

@media print {
    body { background: white; }
    .container { box-shadow: none; margin: 0; padding: 20px; }
    .cards { page-break-inside: avoid; }
    table { page-break-inside: avoid; }
}
</style>
</head>
<body>
<div class="container">

<!-- HEADER -->
<div class="header">
    <h1>{{ info.product }} — Release Report <span class="badge-release">{{ info.release }}</span></h1>
    <p class="subtitle">{{ info.date }} | Prepared by {{ info.prepared_by }} | {{ info.environment }}</p>
</div>

<!-- 1. EXECUTIVE SUMMARY -->
<h2>1. Executive Summary</h2>
<div class="cards">
    <div class="card"><div class="num">{{ stats.total }}</div><div class="label">Total Tests</div></div>
    <div class="card card-pass"><div class="num">{{ stats.passed }}</div><div class="label">Passed</div></div>
    <div class="card card-fail"><div class="num">{{ stats.failed }}</div><div class="label">Failed</div></div>
    <div class="card card-rate"><div class="num">{{ stats.pass_rate }}%</div><div class="label">Pass Rate</div></div>
    <div class="card card-new"><div class="num">{{ stats.new_features }}</div><div class="label">New Features</div></div>
</div>

<p>This document provides formal evidence of testing performed for release <strong>{{ info.release }}</strong> of {{ info.product }}.
Testing covered {{ stats.new_features }} new feature(s), {{ stats.bug_fixes }} bug fix(es), {{ stats.enhancements }} enhancement(s),
and {{ stats.regressions }} regression test(s).</p>

{% if stats.failed == 0 %}
<p style="margin-top:10px;"><strong>Recommendation:</strong> Release {{ info.release }} is approved for promotion from a QA perspective.</p>
{% else %}
<p style="margin-top:10px; color: var(--danger);"><strong>⚠️ {{ stats.failed }} test(s) failed.</strong> Review defect details before release promotion.</p>
{% endif %}

<!-- 2. NEW FEATURES DELIVERED -->
<h2>2. New Features Delivered</h2>
{% if new_features %}
<ul class="feature-list">
{% for f in new_features %}
    <li><span class="feat-module">[{{ f.module }}]</span> <span class="feat-desc">{{ f.feature }}{% if f.description %} — {{ f.description }}{% endif %}</span></li>
{% endfor %}
</ul>
{% else %}
<p>No new features in this release (maintenance/bug-fix release).</p>
{% endif %}

<!-- 3. RELEASE INFORMATION -->
<h2>3. Release Information</h2>
<table class="meta-table">
{% for key, val in info.items() %}
<tr><td>{{ key | replace('_', ' ') | title }}</td><td>{{ val }}</td></tr>
{% endfor %}
</table>

<!-- 4. TEST RESULTS BY MODULE -->
<h2>4. Test Results</h2>
{% for module, tasks in tasks_by_module.items() %}
<h3>{{ module }} <span style="font-size:12px; color:#888;">({{ tasks | length }} tests)</span></h3>
<table>
<thead><tr><th>#</th><th>Task ID</th><th>Feature / Test Case</th><th>Type</th><th>Priority</th><th>Status</th><th>Comments</th></tr></thead>
<tbody>
{% for t in tasks %}
<tr>
    <td>{{ loop.index }}</td>
    <td>{{ t.task_id }}</td>
    <td><strong>{{ t.feature }}</strong>{% if t.description %}<br><small style="color:#666">{{ t.description }}</small>{% endif %}</td>
    <td>{% if t.type == 'New Feature' %}<span class="badge badge-new">New</span>
        {% elif t.type == 'Bug Fix' %}<span class="badge badge-bug">Bug</span>
        {% elif t.type == 'Enhancement' %}<span class="badge badge-enh">Enh</span>
        {% elif t.type == 'Regression' %}<span class="badge badge-reg">Reg</span>
        {% else %}<span class="badge badge-tech">{{ t.type }}</span>{% endif %}</td>
    <td>{% if t.priority == 'P1' %}<span class="badge badge-p1">P1</span>
        {% elif t.priority == 'P2' %}<span class="badge badge-p2">P2</span>
        {% else %}<span class="badge badge-p3">{{ t.priority }}</span>{% endif %}</td>
    <td>{% if t.status.lower() == 'pass' %}<span class="status-pass">✅ Pass</span>
        {% elif t.status.lower() == 'fail' %}<span class="status-fail">❌ Fail</span>
        {% elif t.status.lower() == 'blocked' %}<span class="status-blocked">⚠️ Blocked</span>
        {% else %}{{ t.status }}{% endif %}</td>
    <td>{{ t.comments }}</td>
</tr>
{% endfor %}
</tbody>
</table>
{% endfor %}

<!-- 5. COVERAGE SUMMARY -->
<h2>5. Coverage Summary</h2>
<table>
<thead><tr><th>Module</th><th>Total</th><th>Pass</th><th>Fail</th><th>Blocked</th><th>Pass Rate</th></tr></thead>
<tbody>
{% for module, data in module_stats.items() %}
<tr>
    <td>{{ module }}</td>
    <td>{{ data.total }}</td>
    <td>{{ data.passed }}</td>
    <td>{{ data.failed }}</td>
    <td>{{ data.blocked }}</td>
    <td>{% if data.failed > 0 %}<span class="status-fail">{% else %}<span class="status-pass">{% endif %}{{ data.rate }}%</span></td>
</tr>
{% endfor %}
<tr style="font-weight:700; background:#f5f8fa;">
    <td>TOTAL</td>
    <td>{{ stats.total }}</td>
    <td>{{ stats.passed }}</td>
    <td>{{ stats.failed }}</td>
    <td>{{ stats.blocked }}</td>
    <td>{{ stats.pass_rate }}%</td>
</tr>
</tbody>
</table>

<!-- 6. DEFECTS -->
<h2>6. Defects</h2>
{% if defects %}
<table>
<thead><tr><th>ID</th><th>Severity</th><th>Module</th><th>Description</th><th>Status</th><th>JIRA</th></tr></thead>
<tbody>
{% for d in defects %}
<tr>
    <td>{{ d.defect_id }}</td>
    <td>{{ d.severity }}</td>
    <td>{{ d.module }}</td>
    <td>{{ d.description }}</td>
    <td>{{ d.status }}</td>
    <td>{% if d.jira_link %}<a href="{{ d.jira_link }}">{{ d.jira_link | truncate(30) }}</a>{% endif %}</td>
</tr>
{% endfor %}
</tbody>
</table>
{% else %}
<p class="no-defects">✅ No defects found during this testing cycle.</p>
{% endif %}

<!-- 7. ENVIRONMENT -->
{% if env_config %}
<h2>7. Test Environment</h2>
<table class="meta-table">
{% for row in env_config %}
<tr><td>{{ row.parameter }}</td><td>{{ row.value }}</td></tr>
{% endfor %}
</table>
{% endif %}

<!-- 8. RISKS -->
{% if risks %}
<h2>8. Risks & Assumptions</h2>
<table>
<thead><tr><th>#</th><th>Type</th><th>Description</th><th>Mitigation</th></tr></thead>
<tbody>
{% for r in risks %}
<tr><td>{{ loop.index }}</td><td>{{ r.type }}</td><td>{{ r.description }}</td><td>{{ r.mitigation }}</td></tr>
{% endfor %}
</tbody>
</table>
{% endif %}

<!-- 9. SIGN-OFF -->
<h2>{{ '9' if risks else '8' }}. Sign-Off</h2>
<div class="sign-off">
<table>
<thead><tr><th>Role</th><th>Name</th><th>Date</th><th>Signature</th></tr></thead>
<tbody>
<tr><td>QA Lead</td><td></td><td></td><td></td></tr>
<tr><td>Dev Lead</td><td></td><td></td><td></td></tr>
<tr><td>Product Owner</td><td></td><td></td><td></td></tr>
<tr><td>Release Manager</td><td></td><td></td><td></td></tr>
</tbody>
</table>
</div>

<div class="footer">
    Generated on {{ generated_at }} | {{ info.product }} Release {{ info.release }} | Confidential — Internal Use
</div>

</div>
</body>
</html>"""


# =============================================================================
# EXCEL READING
# =============================================================================
def read_excel(filepath):
    """Read all sheets from the Excel workbook."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- Release Info ---
    info = OrderedDict()
    for name in ["Release Info", "ReleaseInfo", "Summary", "Info"]:
        if name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                if row[0] and row[1]:
                    key = str(row[0]).strip().lower().replace(" ", "_")
                    info[key] = str(row[1]).strip()
            break
    info.setdefault("product", "FacctList")
    info.setdefault("release", "v1.0.0")
    info.setdefault("date", datetime.now().strftime("%B %Y"))
    info.setdefault("prepared_by", "QA Team")
    info.setdefault("environment", "QA")

    # --- Tasks ---
    tasks = []
    for name in ["Tasks", "Test Cases", "TestCases", "Tests", "Sheet1"]:
        if name in wb.sheetnames:
            ws = wb[name]
            headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rec = {headers[i]: str(v).strip() if v else "" for i, v in enumerate(row) if i < len(headers)}
                tasks.append({
                    "task_id": rec.get("task_id", rec.get("id", rec.get("ticket", ""))),
                    "module": rec.get("module", rec.get("category", "General")),
                    "feature": rec.get("feature", rec.get("test_case", rec.get("name", ""))),
                    "description": rec.get("description", rec.get("details", "")),
                    "type": rec.get("type", rec.get("test_type", "Regression")),
                    "priority": rec.get("priority", "P2"),
                    "status": rec.get("status", rec.get("result", "Pass")),
                    "comments": rec.get("comments", rec.get("notes", "")),
                })
            break

    # --- Defects ---
    defects = []
    if "Defects" in wb.sheetnames:
        ws = wb["Defects"]
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rec = {headers[i]: str(v).strip() if v else "" for i, v in enumerate(row) if i < len(headers)}
            defects.append({
                "defect_id": rec.get("defect_id", rec.get("id", "")),
                "severity": rec.get("severity", "Medium"),
                "module": rec.get("module", ""),
                "description": rec.get("description", ""),
                "status": rec.get("status", "Open"),
                "jira_link": rec.get("jira_link", rec.get("jira", rec.get("link", ""))),
            })

    # --- Environment ---
    env_config = []
    if "Environment" in wb.sheetnames:
        ws = wb["Environment"]
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0]:
                env_config.append({"parameter": str(row[0]).strip(), "value": str(row[1] or "").strip()})

    # --- Risks ---
    risks = []
    if "Risks" in wb.sheetnames:
        ws = wb["Risks"]
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rec = {headers[i]: str(v).strip() if v else "" for i, v in enumerate(row) if i < len(headers)}
            risks.append({
                "type": rec.get("type", "Risk"),
                "description": rec.get("description", ""),
                "mitigation": rec.get("mitigation", ""),
            })

    return info, tasks, defects, env_config, risks


# =============================================================================
# REPORT GENERATION
# =============================================================================
def compute_stats(tasks):
    """Compute summary statistics."""
    total = len(tasks)
    passed = sum(1 for t in tasks if t["status"].lower() == "pass")
    failed = sum(1 for t in tasks if t["status"].lower() == "fail")
    blocked = sum(1 for t in tasks if t["status"].lower() == "blocked")
    pass_rate = round((passed / total * 100), 1) if total > 0 else 0
    new_features = sum(1 for t in tasks if t["type"].lower() in ("new feature", "new"))
    bug_fixes = sum(1 for t in tasks if t["type"].lower() in ("bug fix", "bug", "bugfix"))
    enhancements = sum(1 for t in tasks if t["type"].lower() in ("enhancement", "enh"))
    regressions = sum(1 for t in tasks if t["type"].lower() in ("regression", "reg"))

    return {
        "total": total, "passed": passed, "failed": failed, "blocked": blocked,
        "pass_rate": pass_rate, "new_features": new_features,
        "bug_fixes": bug_fixes, "enhancements": enhancements, "regressions": regressions,
    }


def compute_module_stats(tasks):
    """Per-module breakdown."""
    modules = OrderedDict()
    for t in tasks:
        m = t["module"]
        if m not in modules:
            modules[m] = {"total": 0, "passed": 0, "failed": 0, "blocked": 0}
        modules[m]["total"] += 1
        if t["status"].lower() == "pass":
            modules[m]["passed"] += 1
        elif t["status"].lower() == "fail":
            modules[m]["failed"] += 1
        elif t["status"].lower() == "blocked":
            modules[m]["blocked"] += 1

    for m in modules:
        total = modules[m]["total"]
        passed = modules[m]["passed"]
        modules[m]["rate"] = round((passed / total * 100), 1) if total > 0 else 0

    return modules


def generate_html(info, tasks, defects, env_config, risks):
    """Render the HTML report."""
    stats = compute_stats(tasks)
    module_stats = compute_module_stats(tasks)

    # Group tasks by module
    tasks_by_module = OrderedDict()
    for t in tasks:
        m = t["module"]
        if m not in tasks_by_module:
            tasks_by_module[m] = []
        tasks_by_module[m].append(t)

    # Extract new features for the highlights section
    new_features = [t for t in tasks if t["type"].lower() in ("new feature", "new")]

    template = Template(HTML_TEMPLATE)
    return template.render(
        info=info,
        stats=stats,
        module_stats=module_stats,
        tasks_by_module=tasks_by_module,
        new_features=new_features,
        defects=defects,
        env_config=env_config,
        risks=risks,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )


def generate_pdf(html_content, pdf_path):
    """Generate PDF from HTML."""
    try:
        from weasyprint import HTML
        HTML(string=html_content).write_pdf(pdf_path)
        print(f"  PDF:  {pdf_path}")
        return True
    except ImportError:
        pass
    try:
        import pdfkit
        pdfkit.from_string(html_content, pdf_path)
        print(f"  PDF:  {pdf_path}")
        return True
    except ImportError:
        pass
    print("  PDF:  Skipped (install weasyprint or pdfkit)")
    return False


# =============================================================================
# SAMPLE EXCEL TEMPLATE
# =============================================================================
def create_sample_excel(filepath):
    """Create a ready-to-fill Excel template."""
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # --- Release Info ---
    ws = wb.active
    ws.title = "Release Info"
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    data = [
        ("Product", "FacctList"),
        ("Release", "v1.3.0"),
        ("Date", datetime.now().strftime("%B %Y")),
        ("Sprint", "Sprint 24"),
        ("Prepared By", "QA Team"),
        ("Environment", "QA - https://qa-saas.facctum.com"),
        ("Browser", "Chromium (latest)"),
        ("Framework", "Playwright + Cucumber.js (TypeScript)"),
    ]
    for row in data:
        ws.append(row)

    # --- Tasks ---
    ws2 = wb.create_sheet("Tasks")
    headers = ["Task ID", "Module", "Feature", "Description", "Type", "Priority", "Status", "Comments"]
    ws2.append(headers)
    style_header(ws2, len(headers))
    sample_tasks = [
        ("TASK-101", "Suppress/Enrich", "Attribute Suppress - Alias", "Suppress alias with all form fields", "New Feature", "P1", "Pass", ""),
        ("TASK-102", "Suppress/Enrich", "Attribute Enrich - DOB", "Enrich DOB with maker-checker", "New Feature", "P1", "Pass", ""),
        ("TASK-103", "Auto Closure", "Source Update triggers auto-close", "Auto-closure on source delete", "New Feature", "P1", "Pass", ""),
        ("TASK-104", "Concurrency", "Stale tab - version conflict", "POST suppress after approval in another tab", "Regression", "P1", "Pass", "Version conflict detected"),
        ("TASK-105", "Bulk Operations", "Bulk delete multiple records", "Select and delete batch", "Regression", "P2", "Pass", ""),
        ("TASK-106", "API", "Rate limiting 500 req/90s", "No 429 errors observed", "Regression", "P2", "Pass", ""),
        ("TASK-107", "Reporting", "Keyword report export", "Export matches source entries", "Enhancement", "P3", "Pass", ""),
        ("TASK-108", "Attachment", "Download from profile view", "Filename verification", "New Feature", "P2", "Pass", ""),
    ]
    for row in sample_tasks:
        ws2.append(row)
    for col in range(1, 9):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # --- Defects ---
    ws3 = wb.create_sheet("Defects")
    headers = ["Defect ID", "Severity", "Module", "Description", "Status", "JIRA Link"]
    ws3.append(headers)
    style_header(ws3, len(headers))

    # --- Environment ---
    ws4 = wb.create_sheet("Environment")
    headers = ["Parameter", "Value"]
    ws4.append(headers)
    style_header(ws4, len(headers))
    env_data = [
        ("Application URL", "https://qa-saas.facctum.com"),
        ("API URL", "https://qa-api.facctum.com"),
        ("Browser", "Chromium (latest)"),
        ("Viewport", "1920 x 1080"),
        ("Maker User", "reema.singh@facctum.com"),
        ("Approver User", "reema.singh+2@facctum.com"),
        ("Database", "screenDB (MongoDB)"),
    ]
    for row in env_data:
        ws4.append(row)
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 45

    # --- Risks ---
    ws5 = wb.create_sheet("Risks")
    headers = ["Type", "Description", "Mitigation"]
    ws5.append(headers)
    style_header(ws5, len(headers))
    ws5.append(("Assumption", "Test data pre-exists in QA environment", "Verified before execution"))
    ws5.append(("Risk", "Concurrent sessions may behave differently under load", "Covered by stale-tab tests"))
    for col in range(1, 4):
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 35

    wb.save(filepath)
    print(f"✅ Sample Excel template created: {filepath}")
    print(f"   Fill it with your release data and re-run the script.")


# =============================================================================
# MAIN
# =============================================================================
def main():
    parser = argparse.ArgumentParser(description="Generate HTML/PDF release report from Excel")
    parser.add_argument("excel_file", nargs="?", help="Path to Excel workbook")
    parser.add_argument("--output", "-o", help="Output directory (default: same as input)")
    parser.add_argument("--sample", action="store_true", help="Create a sample Excel template")
    args = parser.parse_args()

    if args.sample:
        sample_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "release_report_template.xlsx")
        create_sample_excel(sample_path)
        return

    if not args.excel_file:
        parser.print_help()
        print("\nExample:")
        print("  python generate_release_report.py release_data.xlsx")
        print("  python generate_release_report.py --sample")
        return

    if not os.path.exists(args.excel_file):
        sys.exit(f"ERROR: File not found: {args.excel_file}")

    # Read Excel
    print(f"\n📄 Reading: {args.excel_file}")
    info, tasks, defects, env_config, risks = read_excel(args.excel_file)
    stats = compute_stats(tasks)
    print(f"   Product: {info.get('product')} {info.get('release')}")
    print(f"   Tasks: {len(tasks)} | Defects: {len(defects)} | New Features: {stats['new_features']}")

    # Generate HTML
    html_content = generate_html(info, tasks, defects, env_config, risks)

    # Output paths
    out_dir = args.output or os.path.dirname(os.path.abspath(args.excel_file))
    release_tag = info.get("release", "release").replace(" ", "_")
    base_name = f"{info.get('product', 'Release')}_{release_tag}_report".replace(" ", "_")

    html_path = os.path.join(out_dir, base_name + ".html")
    pdf_path = os.path.join(out_dir, base_name + ".pdf")

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"\n📊 Reports generated:")
    print(f"  HTML: {html_path}")
    generate_pdf(html_content, pdf_path)

    print(f"\n✅ Done! Pass rate: {stats['pass_rate']}% ({stats['passed']}/{stats['total']})")


if __name__ == "__main__":
    main()
