"""
Generate QA Release Report HTML from Excel file: Release 1.23.0 -C2.xlsx
Categorized by Facctum app modules with New Features section.

Usage:
    python scripts/generate_release_report_c2.py
    python scripts/generate_release_report_c2.py --input "path/to/excel.xlsx"
    python scripts/generate_release_report_c2.py --output "path/to/output.html"
"""

import sys
import os
import html
import argparse
from collections import defaultdict, OrderedDict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# --- Configuration ---
DEFAULT_INPUT = os.path.join(os.environ.get("USERPROFILE", ""), "Downloads", "Release 1.23.0 -C2.xlsx")
DEFAULT_OUTPUT = os.path.join(os.environ.get("USERPROFILE", ""), "Downloads", "Release_1.23.0-C2_QA_Report.html")

RELEASE_NAME = "Release 1.23.0 — C2"
REPORT_DATE = "June 2026"
QA_TEAM = "SM Ashwin Kumar, Dipika Bairagi, Madhu P, Sachin Namdev Pawar, Hemant Kumar Pradhan, Reema Singh"

# QA Engineers short name map
QA_NAME_MAP = {
    "SM Ashwin Kumar": "Ashwin",
    "Dipika Bairagi": "Dipika",
    "Madhu P": "Madhu",
    "Sachin Namdev Pawar": "Sachin",
    "Hemant Kumar Pradhan": "Hemant",
    "Reema Singh": "Reema",
}

# ============================================================================
# MANUAL CATEGORIZATION by Facctum App Module
# Tasks are categorized into app features/modules matching the C1 report style
# ============================================================================

TASK_CATEGORIES = OrderedDict([
    ("IBL & Internal List", [4, 8, 11, 31, 59, 63, 72, 74, 76, 77, 99, 101, 112]),
    ("DowJones Suppress/Enrich", [10, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 32, 33, 34, 35, 36, 37, 38, 40, 50, 51, 54, 55, 64, 65, 80, 92, 96, 102, 103, 116]),
    ("Reconciliation", [29, 42, 44, 46, 47, 60, 61, 62, 94, 104]),
    ("WCOD", [1, 95]),
    ("Templates & Export", [3, 28, 39, 48, 79, 89, 91, 97, 110, 118]),
    ("Platform & TMS", [12, 52, 58, 67, 82, 85, 86, 87, 88, 93, 107, 109, 113, 115]),
    ("Regulatory Lists", [7, 13, 56, 68, 73, 100]),
    ("Data & Ingestion", [6, 41, 43, 49, 57, 69, 81, 108]),
    ("Infrastructure & DevOps", [2, 5, 9, 30, 45, 53, 78, 105, 106, 114]),
    ("UI & Task Screen", [66, 70, 71, 75, 83, 84, 98]),
    ("Search & Filters", [111]),
])

# NEW FEATURES delivered in C2 (not in C1 release)
# These are tasks that represent new functionality vs bug fixes from C1
NEW_FEATURES = [
    (11, "Internal list review workflow bulk-service changes"),
    (42, "DB: OFAC vs DJ recon list configuration"),
    (43, "A>I: DJ Data Analysis"),
    (44, "A>I: DJ Recon framework build"),
    (46, "Appropriate Label For Missing Record in DJ Recon Configuration page"),
    (47, "DB: UK Sanctions vs DJ recon list configuration"),
    (48, "Output file format"),
    (54, "Analysis: DB changes for DOB suppress/enrich"),
    (60, "Reconciliation screen redesign"),
    (61, "Add new column parentreconprefid in reconprefref"),
    (62, "Update description for reconlists in reconprefref"),
    (65, "Only DOB should present in Date type while enriching"),
    (67, "Remove 'System Maintenance' notification from TMS platform feature to Platform configuration"),
    (79, "DOB enhancement in Template"),
    (80, "Dowjones idTypes Ref data load"),
    (89, "Barclays template- Counts in template are on row level, Change to record level"),
    (91, "Ab Initio Changes for all 3 templates for Barclays"),
    (96, "Dow Jones: Suppress and enrichment filter and download/Bulk upload"),
    (100, "Sandbox split - Logic to send success and failure mails added for centro plan"),
    (104, "Verify the LSEG buckets for error records - Phase 1"),
    (108, "DB: Update tenantetlconfig and listetlref for dowjones"),
    (110, "BNPP eupr template getting failed"),
    (111, "Fine tune global and records view search by filter the data - Phase 1"),
    (116, "Observations: Suppressed record behavior during update"),
    (118, "Delivery of 2 output templates for BNPP"),
]


def escape(text):
    """HTML-escape text safely."""
    if text is None:
        return "—"
    return html.escape(str(text))


def extract_qa_owners(assignee_str):
    """Extract QA engineer short names from assignee string."""
    if not assignee_str:
        return "—"
    owners = []
    for full_name, short_name in QA_NAME_MAP.items():
        if full_name in str(assignee_str):
            owners.append(short_name)
    return ", ".join(owners) if owners else "—"


def read_sheet(wb, sheet_name):
    """Read all rows from a sheet, returning header + data rows."""
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return rows[0], rows[1:]


def build_qa_coverage(tasks):
    """Build QA coverage summary from tasks."""
    coverage = defaultdict(lambda: {"count": 0, "areas": set()})

    # Build a task_num -> category map
    task_cat_map = {}
    for cat, nums in TASK_CATEGORIES.items():
        for n in nums:
            task_cat_map[n] = cat

    for task in tasks:
        num, name, url, assignee = task
        if not assignee:
            continue
        category = task_cat_map.get(num, "Other")
        for full_name in QA_NAME_MAP:
            if full_name in str(assignee):
                coverage[full_name]["count"] += 1
                coverage[full_name]["areas"].add(category)
    return coverage


def generate_html(input_file, output_file):
    """Generate the HTML report from Excel data."""
    print(f"Reading: {input_file}")
    wb = openpyxl.load_workbook(input_file, read_only=True)

    # Read all sheets
    _, tasks = read_sheet(wb, "Tasks")
    _, spillover = read_sheet(wb, "QA Spillover")
    _, open_bugs = read_sheet(wb, "Open Bugs for C3")
    _, closed_bugs = read_sheet(wb, "Closed Bugs")

    # Build task lookup by number
    task_map = {}
    for task in tasks:
        num, name, url, assignee = task
        if num is not None:
            task_map[num] = task

    # Counts
    total_tasks = len(tasks)
    total_spillover = len(spillover)
    total_open_bugs = len(open_bugs)
    total_closed_bugs = len(closed_bugs)

    # QA Coverage
    coverage = build_qa_coverage(tasks)

    # Start building HTML
    html_parts = []
    html_parts.append(generate_header())
    html_parts.append(generate_summary(total_tasks, total_closed_bugs, total_spillover, total_open_bugs))
    html_parts.append(generate_new_features_section())
    html_parts.append('<hr class="section-divider">')
    html_parts.append(generate_qa_coverage(coverage))
    html_parts.append('<hr class="section-divider">')
    html_parts.append(generate_tasks_section(task_map, total_tasks))
    html_parts.append('<hr class="section-divider">')
    html_parts.append(generate_closed_bugs_section(closed_bugs, total_closed_bugs))
    html_parts.append('<hr class="section-divider">')
    html_parts.append(generate_spillover_section(spillover, total_spillover))
    html_parts.append('<hr class="section-divider">')
    html_parts.append(generate_open_bugs_section(open_bugs, total_open_bugs))
    html_parts.append('<hr class="section-divider">')
    html_parts.append(generate_recommendation(total_tasks, total_closed_bugs, total_open_bugs, total_spillover))
    html_parts.append(generate_signoff())
    html_parts.append(generate_footer())

    report_html = "\n".join(html_parts)

    # Write output
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(report_html)

    print(f"Report generated: {output_file}")
    print(f"  Tasks: {total_tasks} | Closed Bugs: {total_closed_bugs} | Spillover: {total_spillover} | Open Bugs: {total_open_bugs}")
    print(f"  New Features in C2: {len(NEW_FEATURES)}")


def generate_header():
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>QA Release Report — {escape(RELEASE_NAME)}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #f4f6f9; color: #333; line-height: 1.6; padding: 24px; }}
  .container {{ max-width: 1200px; margin: 0 auto; background: #fff; border-radius: 8px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); padding: 40px; }}
  h1 {{ color: #1a237e; border-bottom: 3px solid #1a237e; padding-bottom: 12px; margin-bottom: 8px; font-size: 28px; }}
  .subtitle {{ color: #555; margin-bottom: 24px; font-size: 14px; }}
  h2 {{ color: #283593; margin-top: 36px; margin-bottom: 12px; font-size: 20px; border-left: 4px solid #3949ab; padding-left: 12px; }}
  h3 {{ color: #37474f; margin-top: 24px; margin-bottom: 8px; font-size: 16px; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0 24px 0; font-size: 13px; }}
  th {{ background: #1a237e; color: #fff; padding: 10px 12px; text-align: left; font-weight: 600; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #e0e0e0; }}
  tr:nth-child(even) {{ background: #f8f9fc; }}
  tr:hover {{ background: #e8eaf6; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
  .badge-green {{ background: #c8e6c9; color: #2e7d32; }}
  .badge-red {{ background: #ffcdd2; color: #c62828; }}
  .badge-orange {{ background: #fff3e0; color: #e65100; }}
  .badge-blue {{ background: #bbdefb; color: #1565c0; }}
  .badge-purple {{ background: #e1bee7; color: #6a1b9a; }}
  .badge-gray {{ background: #eceff1; color: #546e7f; }}
  .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 16px; margin: 16px 0 24px 0; }}
  .summary-card {{ background: #f8f9fc; border: 1px solid #e0e0e0; border-radius: 8px; padding: 16px; text-align: center; }}
  .summary-card .number {{ font-size: 32px; font-weight: 700; color: #1a237e; }}
  .summary-card .label {{ font-size: 12px; color: #666; margin-top: 4px; }}
  .summary-card.highlight {{ border-color: #4caf50; background: #e8f5e9; }}
  .summary-card.highlight .number {{ color: #2e7d32; }}
  .new-features {{ background: #e3f2fd; border: 2px solid #1976d2; border-radius: 8px; padding: 20px; margin: 24px 0; }}
  .new-features h3 {{ color: #1565c0; margin-top: 0; margin-bottom: 12px; }}
  .recommendation {{ background: #e8f5e9; border: 2px solid #4caf50; border-radius: 8px; padding: 20px; margin: 24px 0; }}
  .recommendation h3 {{ color: #2e7d32; margin-top: 0; }}
  .recommendation ul {{ margin: 8px 0 0 20px; }}
  .signoff-table td:last-child {{ text-align: center; font-size: 18px; }}
  .section-divider {{ border: none; border-top: 2px solid #e0e0e0; margin: 32px 0; }}
  .risk-high {{ color: #c62828; font-weight: 700; }}
  .risk-medium {{ color: #e65100; font-weight: 600; }}
  .risk-low {{ color: #2e7d32; }}
  a {{ color: #1565c0; text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  @media print {{ body {{ background: #fff; padding: 0; }} .container {{ box-shadow: none; padding: 20px; }} }}
</style>
</head>
<body>
<div class="container">

<h1>QA Release Report — {escape(RELEASE_NAME)}</h1>
<div class="subtitle">
  <strong>Prepared by:</strong> QA Team &nbsp;|&nbsp;
  <strong>Date:</strong> {REPORT_DATE} &nbsp;|&nbsp;
  <strong>QA Team:</strong> {escape(QA_TEAM)}
</div>"""


def generate_summary(tasks, closed, spillover, open_bugs):
    return f"""
<h2>Executive Summary</h2>
<div class="summary-grid">
  <div class="summary-card"><div class="number">{tasks}</div><div class="label">Total Tasks</div></div>
  <div class="summary-card highlight"><div class="number">{len(NEW_FEATURES)}</div><div class="label">New Features (C2)</div></div>
  <div class="summary-card"><div class="number">{closed}</div><div class="label">Closed Bugs</div></div>
  <div class="summary-card"><div class="number">{spillover}</div><div class="label">QA Spillover</div></div>
  <div class="summary-card"><div class="number">{open_bugs}</div><div class="label">Open Bugs (C3)</div></div>
</div>"""


def generate_new_features_section():
    """Generate highlighted section for new features delivered in C2."""
    lines = []
    lines.append("""
<div class="new-features">
  <h3>🆕 New Features Delivered in C2</h3>
  <p style="margin-bottom:12px; color:#555;">The following new features/enhancements were delivered in C2 (not part of C1 release):</p>
  <table>
    <tr><th>#</th><th>Feature</th></tr>""")

    for num, desc in NEW_FEATURES:
        lines.append(f'    <tr><td>{num}</td><td>{escape(desc)}</td></tr>')

    lines.append("""  </table>
</div>""")
    return "\n".join(lines)


def generate_qa_coverage(coverage):
    """Generate QA coverage table."""
    lines = ['<h2>QA Coverage — Testing Ownership</h2>', '<table>',
             '  <tr><th>QA Engineer</th><th>Tasks Assigned</th><th>Key Areas</th></tr>']

    sorted_coverage = sorted(coverage.items(), key=lambda x: x[1]["count"], reverse=True)
    for name, data in sorted_coverage:
        areas = ", ".join(sorted(data["areas"]))
        lines.append(f'  <tr><td><strong>{escape(name)}</strong></td><td>{data["count"]}</td><td>{escape(areas)}</td></tr>')

    lines.append('</table>')
    return "\n".join(lines)


def generate_tasks_section(task_map, total):
    """Generate categorized tasks section by Facctum app module."""
    lines = [f'<h2>Tasks Completed ({total})</h2>']

    # Track uncategorized tasks
    categorized_nums = set()
    for nums in TASK_CATEGORIES.values():
        categorized_nums.update(nums)

    for category, task_nums in TASK_CATEGORIES.items():
        # Filter to tasks that exist in the Excel
        existing = [(n, task_map[n]) for n in task_nums if n in task_map]
        if not existing:
            continue

        lines.append(f'\n<h3>{escape(category)} ({len(existing)})</h3>')
        lines.append('<table>')
        lines.append('  <tr><th>#</th><th>Task</th><th>QA Owner</th></tr>')

        for num, task in existing:
            _, name, url, assignee = task
            name_str = escape(str(name)) if name else "—"
            if url:
                name_str = f'<a href="{escape(str(url))}" target="_blank">{name_str}</a>'
            qa_owners = extract_qa_owners(assignee)
            lines.append(f'  <tr><td>{num}</td><td>{name_str}</td><td>{qa_owners}</td></tr>')

        lines.append('</table>')

    # Uncategorized tasks
    uncategorized = [(n, task_map[n]) for n in sorted(task_map.keys()) if n not in categorized_nums]
    if uncategorized:
        lines.append(f'\n<h3>Other ({len(uncategorized)})</h3>')
        lines.append('<table>')
        lines.append('  <tr><th>#</th><th>Task</th><th>QA Owner</th></tr>')
        for num, task in uncategorized:
            _, name, url, assignee = task
            name_str = escape(str(name)) if name else "—"
            if url:
                name_str = f'<a href="{escape(str(url))}" target="_blank">{name_str}</a>'
            qa_owners = extract_qa_owners(assignee)
            lines.append(f'  <tr><td>{num}</td><td>{name_str}</td><td>{qa_owners}</td></tr>')
        lines.append('</table>')

    return "\n".join(lines)


def generate_closed_bugs_section(closed_bugs, total):
    """Generate closed bugs section grouped by area."""
    lines = [f'<h2>Closed Bugs ({total})</h2>']

    # Sub-categorize closed bugs
    dj_se_bugs = []
    ibl_bugs = []
    template_bugs = []
    platform_bugs = []
    recon_bugs = []
    other_bugs = []

    for bug in closed_bugs:
        num, desc, url, assignee = bug
        desc_str = str(desc).lower() if desc else ""

        if any(kw in desc_str for kw in ["suppress", "enrich", "dowjone", "dj ", "maiden", "original script",
                                          "dob ", "date type", "delta", "o/p file", "audit not capturing",
                                          "next button", "surname", "id notes", "version is missing"]):
            dj_se_bugs.append(bug)
        elif any(kw in desc_str for kw in ["ibl", "internal list", "bulk"]):
            ibl_bugs.append(bug)
        elif any(kw in desc_str for kw in ["template", "sftp", "output file", "export"]):
            template_bugs.append(bug)
        elif any(kw in desc_str for kw in ["tms", "role", "webhook", "notification", "sso", "maintenance"]):
            platform_bugs.append(bug)
        elif any(kw in desc_str for kw in ["recon", "minfi", "accuity", "lseg", "mismatch"]):
            recon_bugs.append(bug)
        else:
            other_bugs.append(bug)

    def render_bug_table(title, bugs):
        if not bugs:
            return ""
        result = [f'<h3>{escape(title)} ({len(bugs)})</h3>', '<table>',
                  '  <tr><th>#</th><th>Bug</th><th>QA Verified</th></tr>']
        for bug in bugs:
            num, desc, url, assignee = bug
            num_str = str(num) if num else "—"
            desc_str = escape(str(desc)) if desc else "—"
            if url:
                desc_str = f'<a href="{escape(str(url))}" target="_blank">{desc_str}</a>'
            # Security badge
            if any(kw in str(desc).lower() for kw in ["security", "unauthorised", "unauthorized"]):
                desc_str = f'<span class="badge badge-red">SECURITY</span> {desc_str}'
            qa = extract_qa_owners(assignee)
            result.append(f'  <tr><td>{escape(num_str)}</td><td>{desc_str}</td><td>{qa}</td></tr>')
        result.append('</table>')
        return "\n".join(result)

    lines.append(render_bug_table("DowJones Suppress/Enrich Bugs", dj_se_bugs))
    lines.append(render_bug_table("IBL & Internal List Bugs", ibl_bugs))
    lines.append(render_bug_table("Template & Export Bugs", template_bugs))
    lines.append(render_bug_table("Platform & TMS Bugs", platform_bugs))
    lines.append(render_bug_table("Reconciliation & Data Bugs", recon_bugs))
    lines.append(render_bug_table("Other Bugs", other_bugs))

    return "\n".join(lines)


def generate_spillover_section(spillover, total):
    """Generate spillover section."""
    lines = [f'<h2>QA Spillover ({total})</h2>']
    lines.append('<table>')
    lines.append('  <tr><th>#</th><th>Task</th><th>Comment / Status</th></tr>')

    for item in spillover:
        num, name, url, comment = item
        num_str = str(num) if num else "—"
        name_str = escape(str(name)) if name else "—"
        if url:
            name_str = f'<a href="{escape(str(url))}" target="_blank">{name_str}</a>'
        comment_str = escape(str(comment)) if comment else "—"
        lines.append(f'  <tr><td>{escape(num_str)}</td><td>{name_str}</td><td><span class="badge badge-orange">{comment_str}</span></td></tr>')

    lines.append('</table>')
    return "\n".join(lines)


def generate_open_bugs_section(open_bugs, total):
    """Generate open bugs section with risk assessment."""
    lines = [f'<h2>Open Bugs for C3 ({total}) — Risk Assessment</h2>']
    lines.append('<table>')
    lines.append('  <tr><th>#</th><th>Bug Description</th><th>Assignee</th><th>Risk</th></tr>')

    for bug in open_bugs:
        num, desc, url, assignee = bug
        num_str = str(num) if num else "—"
        desc_str = escape(str(desc)) if desc else "—"
        if url:
            desc_str = f'<a href="{escape(str(url))}" target="_blank">{desc_str}</a>'
        assignee_str = extract_qa_owners(assignee)
        if assignee_str == "—":
            assignee_str = escape(str(assignee)) if assignee else "—"

        # Assign risk level
        desc_lower = str(desc).lower() if desc else ""
        if any(kw in desc_lower for kw in ["security", "data loss", "crash", "failing", "not working",
                                            "terminated", "failed"]):
            risk = '<span class="risk-high">HIGH</span>'
        elif any(kw in desc_lower for kw in ["incorrect", "missing", "issue", "error", "limit", "not show"]):
            risk = '<span class="risk-medium">MEDIUM</span>'
        else:
            risk = '<span class="risk-low">LOW</span>'

        lines.append(f'  <tr><td>{escape(num_str)}</td><td>{desc_str}</td><td>{assignee_str}</td><td>{risk}</td></tr>')

    lines.append('</table>')
    return "\n".join(lines)


def generate_recommendation(tasks, closed, open_bugs, spillover):
    """Generate release recommendation."""
    return f"""
<div class="recommendation">
  <h3>Release Recommendation: Conditional GO ✅</h3>
  <p><strong>Rationale:</strong></p>
  <ul>
    <li>{tasks} tasks completed and verified in C2</li>
    <li>{len(NEW_FEATURES)} new features/enhancements delivered</li>
    <li>{closed} bugs closed in this cycle</li>
    <li>{spillover} items in QA spillover (tracked for C3)</li>
    <li>{open_bugs} open bugs carried forward to C3</li>
  </ul>
  <p style="margin-top:12px"><strong>Conditions for Release:</strong></p>
  <ul>
    <li>High-risk open bugs must be resolved or documented as known issues before production</li>
    <li>DowJones Suppress/Enrich — core workflow validated; remaining edge cases tracked</li>
    <li>IBL builder limit issues (bugs #1-3) should be resolved before production</li>
    <li>QA spillover items to be prioritized in C3</li>
  </ul>
</div>"""


def generate_signoff():
    """Generate sign-off section."""
    lines = ['<h2>Sign-off</h2>', '<table class="signoff-table">',
             '  <tr><th>QA Engineer</th><th>Area Verified</th><th>Sign-off</th></tr>']

    signoff_data = [
        ("SM Ashwin Kumar", "Platform & TMS, Webhooks, IBL Permissions, Templates, Security"),
        ("Dipika Bairagi", "WCOD, Regulatory Lists, Reports, Advance Filters, Ingestion"),
        ("Madhu P", "DowJones S/E, IBL, WC Premium, Reconciliation, UI, WCOD"),
        ("Sachin Namdev Pawar", "IBL, Internal List, Templates, Bulk Workflows, Task View"),
        ("Hemant Kumar Pradhan", "DowJones S/E, Templates, DOB Enhancement, Reconciliation, Audit"),
        ("Reema Singh", "Regulatory Lists, Reconciliation, Templates, Sandbox Split, Barclays"),
    ]

    for name, areas in signoff_data:
        lines.append(f'  <tr><td><strong>{escape(name)}</strong></td><td>{escape(areas)}</td><td>☐</td></tr>')

    lines.append('</table>')
    return "\n".join(lines)


def generate_footer():
    return """
</div>
</body>
</html>"""


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate QA Release Report HTML from Excel")
    parser.add_argument("--input", "-i", default=DEFAULT_INPUT, help="Path to Excel file")
    parser.add_argument("--output", "-o", default=DEFAULT_OUTPUT, help="Output HTML file path")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        print(f"Provide the correct path with: python {sys.argv[0]} --input \"path/to/file.xlsx\"")
        sys.exit(1)

    generate_html(args.input, args.output)
