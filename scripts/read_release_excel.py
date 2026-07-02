import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\ReemaSingh\Downloads\Release 1.23.0 and C1.xlsx', read_only=True)
for s in wb.sheetnames:
    ws = wb[s]
    print(f'=== {s} ===')
    for row in ws.iter_rows(values_only=True):
        print(f'  {row}')
    print()
